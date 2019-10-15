#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019/8/25 19:52
# @Author  : jarvis
# @File    : views.py
# @Software: PyCharm
# @Contact : 309194437@qq.com
"""接口实现"""
import ast
import datetime
import hashlib
import os

import openpyxl as openpyxl
from flask import jsonify, request, json, Response
from flask_restful import Resource, Api, marshal
# from flask_restful import reqparse # 用于请求的参数解析
# from sqlalchemy.orm import session
from openpyxl.styles import Alignment, Font

from api.v1.models import app, Role, User, Task, Department, Product,Dictitem
from api.v1.serializers import *

api = Api(app)

"""格式化返回结果"""


def return_true_json(data):
    return jsonify({
        "status": 1,
        "data": data,
        "msg": "request successfully"
    })

def return_page_true_json(data,page,pages,per_page,has_prev,has_next,total):
    return jsonify({
        "status": 1,
        "data": data,
        "page":page,
        "pages":pages,
        "per_page":per_page,
        "has_prev":has_prev,
        "has_next":has_next,
        "total":total,
        "msg": "request successfully"
    })


def return_false_json(data):
    return jsonify({
        "status": 0,
        "data": data,
        "msg": "request failed"
    })

def return_page_false_json(data,page,pages,per_page,has_prev,has_next,total):
    return jsonify({
        "status": 0,
        "data": data,
        "page":page,
        "pages":pages,
        "per_page":per_page,
        "has_prev":has_prev,
        "has_next":has_next,
        "total":total,
        "msg": "request failed"
    })


# #获取列表里面的字典的值方法(已改成下面的优化方法)
# def get_dict_name(list,key):
#     for i in list:
#         # print(i,type(i),i["key"])
#         if i["key"] == key:
#             return i["value"]
#             break


#获取相应字段数值的字典值
def get_dict_name(dict_name,key):
    dict_value = Dictitem.query.filter(Dictitem.dict_code == dict_name).first().key_value
    # print(dict_value)
    l_dict_value = ast.literal_eval(dict_value)
    # print(l_dict_value)
    for i in l_dict_value:
        if i["key"] == key:
            return i["value"]
            break

# 日期格式转换，将string格式日期转换成datatime格式
def str_to_datatime(time_str):
    time_list = time_str[0:10].split("-")
    year = time_list[0]
    month = time_list[1]
    day = time_list[2]
    return datetime.date(int(year), int(month), int(day))


# 计算逾期天数
def get_diviation(planfinished_time, finished_time):
    if finished_time:
        diviation = (str_to_datatime(finished_time) - str_to_datatime(planfinished_time)).days
        return diviation
    else:
        return None

#根据日期计算当前属于当月第几周,周一作为一周的开始
def get_week_of_month(time_str):
    time_list = time_str[0:10].split("-")
    year = int(time_list[0])
    month = int(time_list[1])
    day = int(time_list[2])
    end = int(datetime.datetime(year, month, day).strftime("%W"))
    begin = int(datetime.datetime(year, month, 1).strftime("%W"))
    return end - begin + 1

#根据传入的参数导出生产EXCEL文件
def creation_excel_xlsx(path, sheet_name, value):
        index = len(value)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
        workbook.save(path)
        # print("xlsx格式表格写入数据成功！")

#追加数据到excel文件中
def write_excel_xlsx(path,sheet_name,value):
    path = path
    wb = openpyxl.load_workbook(path)
    #切换到目标数据表
    ws = wb[sheet_name]
    #设置单元格格式
    font1 = Font(name=u'微软雅黑',size = 11,bold = False,italic = False,vertAlign = None,underline = 'none',strike = False,#设置字体颜色
    color = 'FF000000')
    aligmentCenter = Alignment(horizontal='center', vertical ='center',wrap_text = True,)#文字居中
    #待填充数据
    index = len(value)
    for i in range(0,index):
        ws.append(value[i])
    for eachCommonRow in ws.iter_rows(min_row=3, max_col=ws.max_column, max_row=ws.max_row):
        for eachCellInRow in eachCommonRow:
            eachCellInRow.alignment = aligmentCenter
            eachCellInRow.font = font1
    savename = path
    wb.save(savename)
    # print("xlsx格式表格追加数据成功")


#文件读取迭代器
def file_iterator(file_path, chunk_size=512):
    """
    file_path:文件路径
    chunk_size: 每次读取流大小
    """
    with open(file_path, 'rb') as target_file:
        while True:
            chunk = target_file.read(chunk_size)
            # print(chunk)
            if chunk:
                yield chunk
            else:
                break

'''密码MD5加密方法'''
def get_md5(password):
    obj = hashlib.md5(password.encode('utf-8'))
    obj.update
    result = obj.hexdigest()
    return result
    # print(result)

"""定义资源"""

class Hello(Resource):
    def get(self):
        return 'Hello Flask-restful!'


class Login(Resource):
    def post(self):
        args = parser_login.parse_args()
        account = User.query_account(args['account'])
        # res = [marshal(account, resource_user_fields)]
        print(type(account))
        password = args['password']
        # print(get_md5(password))
        # print(account)
        # print(password)
        if args['account'] and password:
            if account and get_md5(password) == account.password:
                # return return_true_json("登录成功")
                return jsonify({
                    "status": 1,
                    "data": "登录成功",
                    "user_id":account.user_id,
                    "user_name":account.user_name,
                    "msg": "request successfully"
                })
            else:
                return return_false_json("用户名密码错误")
        else:
            return return_false_json("用户名密码不能为空")

# 任务列表接口
class Tasklist(Resource):
    def post(self):
        args = parser_task.parse_args()
        task_type = args['task_type']
        task_nature = args['task_nature']
        pdt_id = args['pdt_id']
        content = args['content']
        planstart_time = args['planstart_time']
        start_time = args['start_time']
        planfinished_time = args['planfinished_time']
        finished_percent = args['finished_percent']
        #如果提交的任务进度是100，则自动更新当前时间为完成时间
        if finished_percent == "100" :
            finished_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        else:
            finished_time = None
        remark = args['remark']
        creat_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        user_id = args['user_id']
        if user_id:     #判断user_id是否为空，为空不能插入
            # print(planfinished_time)
            # print(type(planfinished_time))
            task = Task(task_type=task_type,
                        task_nature=task_nature,
                        pdt_id=pdt_id,
                        content=content,
                        planstart_time=planstart_time,
                        start_time=start_time,
                        planfinished_time=planfinished_time,
                        finished_percent=finished_percent,
                        finished_time = finished_time,
                        create_time=creat_time,
                        user_id=user_id,
                        remark=remark)
            task.add_to_db()
            if (task.task_id is None):
                return return_false_json("任务插入失败")
            else:
                return return_true_json("任务新增成功")
        else:
            return return_false_json("user_id不能为空")

    def put(self):
        args = parser_task.parse_args()
        task_id = args['task_id']
        task_type = args['task_type']
        task_nature = args['task_nature']
        pdt_id = args['pdt_id']
        content = args['content']
        planstart_time = args['planstart_time']
        start_time = args['start_time']
        planfinished_time = args['planfinished_time']
        finished_percent = args['finished_percent']
        # 如果提交的任务进度是100，则自动更新当前时间为完成时间
        if finished_percent == "100" :
            finished_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        else:
            finished_time = None
        user_id = args['user_id']
        remark = args['remark']

        task = Task.query_task_id(task_id)
        task.task_type = task_type
        task.task_nature = task_nature
        task.pdt_id = pdt_id
        task.content = content
        task.planstart_time = planstart_time
        task.start_time = start_time
        task.planfinished_time = planfinished_time
        task.finished_percent = finished_percent
        task.finished_time = finished_time
        task.user_id = user_id
        task.remark = remark
        Task.commit(self)
        return return_true_json("任务更新成功")

    def get(self):
        args = parser_task.parse_args()
        page = args['page']
        per_page = args['per_page']
        user_id = args['user_id']
        planstart_time1 = args['planstart_time1']
        planstart_time2 = args['planstart_time2']
        # datas = Task.find_all()
        if user_id:   #根据user_id和planstart_time是否为空使用相应的方法返回数据
            if planstart_time1:
                paginates = Task.find_self_weekly(page, per_page,user_id,planstart_time1,planstart_time2)
            else:
                paginates = Task.find_self(page, per_page,user_id)
        else:
            paginates = Task.find_all(page, per_page)
        datas = paginates.items
        page = paginates.page #返回第几页
        pages = paginates.pages  #总页数
        per_page = paginates.per_page #每页数量
        has_prev = paginates.has_prev
        has_next = paginates.has_next
        total = paginates.total   #总记录数
        # rwlx_value = Dictitem.query.filter(Dictitem.dict_code == 'rwlx').first().key_value
        # rwxz_value = Dictitem.query.filter(Dictitem.dict_code == 'rwxz').first().key_value
        # print(datas)
        # print(type(datas))
        # print(rwlx_value, type(rwlx_value))
        # l_rwlx_value = ast.literal_eval(rwlx_value)
        # l_rwxz_value = ast.literal_eval(rwxz_value)
        # print(l_rwxz_value,type(l_rwxz_value))
        # 合并字典（key值不能相同）
        # d_rwlx_value = {}
        # for i in l_rwlx_value:
        #     d_rwlx_value = {**d_rwlx_value,**i}
        # print(d_rwlx_value, type(d_rwlx_value))
        als = []
        #把user_name字段加入到task的返回数据中
        for i in range(len(datas)):
            to_json = {'task_id': datas[i].task_id,
                       'content': datas[i].content,
                       'task_type': datas[i].task_type,
                       'task_nature':datas[i].task_nature,
                       'task_nature_name':get_dict_name("rwxz", datas[i].task_nature),
                       'task_type_name':get_dict_name("rwlx", datas[i].task_type),
                       'pdt_id': datas[i].pdt_id,
                       'product_name': Product.find_by_id(datas[i].pdt_id).product_name,
                       'planfinished_time': datas[i].planfinished_time,
                       'finished_time': datas[i].finished_time,
                       'finished_percent': datas[i].finished_percent,
                       'create_time': datas[i].create_time,
                       'user_id': datas[i].user_id,
                       'remark': datas[i].remark,
                       'deviation': get_diviation(datas[i].planfinished_time, datas[i].finished_time),
                       'delete_flag': datas[i].delete_flag,
                       'user_name': Task.query_user_name(datas[i].user_id).user_name
                       }
            als.append(to_json)
        # return als

        res = [marshal(al, resource_task_fields) for al in als]
        if res:
            return return_page_true_json(res,page,pages,per_page,has_prev,has_next,total)
        else:
            return return_page_false_json(res,page,pages,per_page,has_prev,has_next,total)

    def delete(self):
        args = parser_task.parse_args()
        task_id = args['task_id']
        delete_flag = args['delete_flag']

        task = Task.query_task_id(task_id)
        task.delete_flag = delete_flag
        Task.commit(self)
        return return_true_json("删除成功")

#条件查询接口
class QueryTasklist(Resource):
    def post(self):
        args = parser_task.parse_args()
        print(args)
        page = args['page']
        per_page = args['per_page']
        # 创建一个filter的条件列表，将选择的条件加入列表中
        filter = []
        #统计条件支持多选
        if ('user_name' in args) and (args['user_name']!=None):#[None]的布尔值为true]
            user_name = args['user_name']
            print(user_name)
            filter.append(User.user_name.in_(user_name))
        if ('task_type' in args) and (args['task_type']!=None):
            task_type = args['task_type']
            print(task_type)
            filter.append(Task.task_type.in_(task_type))
            # print(filter)
        if ('pdt_id' in args) and (args['pdt_id']!=None):
            pdt_id = args['pdt_id']
            print(pdt_id)
            filter.append(Product.pdt_id.in_(pdt_id))
        # if ('user_name' in args) and (args['user_name']):#[None]的布尔值为true]
        #     user_name = args['user_name']
        #     # print(user_name)
        #     filter.append(User.user_name==user_name)
        # if ('task_type' in args) and (args['task_type']):
        #     task_type = args['task_type']
        #     # print(task_type)
        #     filter.append(Task.task_type==task_type)
        #     print(filter)
        # if ('pdt_id' in args) and (args['pdt_id']):
        #     pdt_id = args['pdt_id']
        #     # print(pdt_id)
        #     filter.append(Product.pdt_id==pdt_id)
        if ('startDate' in request.args) and (request.args['startDate']):
            startDate = request.args['startDate']
            # filter.append(db.cast(Task.finished_time, db.DATE) >= db.cast(startDate, db.Date))
            filter.append(Task.finished_time>= startDate)
        if ('endDate' in request.args) and (request.args['endDate']):
            endDate = request.args['endDate']
            # filter.append(db.cast(Task.finished_time, db.DATE) <= db.cast(endDate, db.Date))
            filter.append(Task.finished_time<= endDate)
        paginates = Task.query.order_by('task_id').filter(Task.user_id==User.user_id).filter(Task.pdt_id==Product.pdt_id).filter(Task.delete_flag==0).filter(*filter).paginate(page, per_page, error_out=False)
        print(filter)
        # datas = Task.query.join(User,Task.user_id==User.user_id).filter(Task.delete_flag==0).filter(*filter).all()
        datas = paginates.items
        page = paginates.page
        pages = paginates.pages
        per_page = paginates.per_page
        has_prev = paginates.has_prev
        has_next = paginates.has_next
        total = paginates.total
        # rwlx_value = Dictitem.query.filter(Dictitem.dict_code == 'rwlx').first().key_value
        # rwxz_value = Dictitem.query.filter(Dictitem.dict_code == 'rwxz').first().key_value
        # l_rwlx_value = ast.literal_eval(rwlx_value)
        # l_rwxz_value = ast.literal_eval(rwxz_value)
        als = []
        # 把user_name等字段加入到datas的返回数据中
        for i in range(len(datas)):
            to_json = {'task_id': datas[i].task_id,
                       'content': datas[i].content,
                       'task_type': datas[i].task_type,
                       'task_nature': datas[i].task_nature,
                       'task_nature_name': get_dict_name("rwxz", datas[i].task_nature),
                       'task_type_name': get_dict_name("rwlx", datas[i].task_type),
                       'pdt_id': datas[i].pdt_id,
                       'product_name': Product.find_by_id(datas[i].pdt_id).product_name,
                       'planfinished_time': datas[i].planfinished_time,
                       'finished_time': datas[i].finished_time,
                       'finished_percent': datas[i].finished_percent,
                       'create_time': datas[i].create_time,
                       'user_id': datas[i].user_id,
                       'remark': datas[i].remark,
                       'deviation': get_diviation(datas[i].planfinished_time, datas[i].finished_time),  #计算延期天数
                       'delete_flag': datas[i].delete_flag,
                       'user_name': Task.query_user_name(datas[i].user_id).user_name
                       }
            als.append(to_json)
        als = [marshal(al, resource_task_fields) for al in als]
        if als:
            return return_page_true_json(als, page, pages, per_page, has_prev, has_next, total)
        else:
            return return_page_false_json(als, page, pages, per_page, has_prev, has_next, total)

class UserList(Resource):
    def post(self):
        args = parser_user.parse_args()
        account = args['account']
        password = args['password']
        user_name = args['user_name']
        create_account = args['create_account']
        create_time = args['create_time']
        role_id = args['role_id']
        dept_id = args['dept_id']
        status = args['status']
        delete_flag = args['delete_flag']

        user = User(account=account,
                    password=password,
                    user_name=user_name,
                    create_account=create_account,
                    create_time=create_time,
                    role_id=role_id,
                    dept_id=dept_id,
                    status=status,
                    delete_flag=delete_flag)
        print(user.user_name)
        user.add_to_db()
        if (user.user_id is None):
            return return_false_json("插入失败")
        else:
            return return_true_json("新增成功")

    def get(self):
        datas = User.find_all()
        res = [marshal(data, resource_user_fields) for data in datas]
        if datas:
            return return_true_json(res)
        else:
            return return_false_json(res)

    def put(self):
        args = parser_user.parse_args()
        user_id = args['user_id']
        account = args['account']
        # password = args['password']
        user_name = args['user_name']
        role_id = args['role_id']
        dept_id = args['dept_id']
        status = args['status']

        user = User.query_user_id(user_id)
        user.account = account
        # user.password = password
        user.user_name = user_name
        user.role_id = role_id
        user.dept_id = dept_id
        user.status = status
        User.commit(self)
        return return_true_json("用户更新成功")

    def delete(self):
        args = parser_user.parse_args()
        user_id = args['user_id']
        delete_flag = args['delete_flag']

        user = User.query_user_id(user_id)
        user.delete_flag = delete_flag
        User.commit(self)
        return return_true_json("用户删除成功")


class RoleList(Resource):
    def get(self):
        args = parser_role.parse_args()
        page = args['page']
        per_page = args['per_page']
        # datas = Role.find_all()
        paginates = Role.find_all(page,per_page)
        datas = paginates.items
        res = [marshal(data, resource_role_fields) for data in datas]
        if datas:
            return return_true_json(res)
        else:
            return return_false_json(res)

    def post(self):
        args = parser_role.parse_args()
        role_name = args['role_name']
        role = Role(role_name=role_name)
        role.add_to_db()
        return return_true_json("新增成功")

    def put(self):
        args = parser_role.parse_args()
        role_id = args['role_id']
        role_name = args['role_name']

        role = Role.find_by_id(role_id)
        role.role_name = role_name
        Role.commit(self)
        return return_true_json("角色更新成功")

    def delete(self):
        args = parser_role.parse_args()
        role_id = args['role_id']

        role = Role.find_by_id(role_id)
        Role.delete(role)
        return return_true_json("删除成功")


class DepartmentList(Resource):
    def get(self):
        datas = Department.find_all()
        res = [marshal(data, resource_department_fields) for data in datas]
        if datas:
            return return_true_json(res)
        else:
            return return_false_json(res)

    def post(self):
        args = parser_department.parse_args()
        department_name = args['department_name']
        department = Department(department_name=department_name)
        department.add_to_db()
        return return_true_json("部门新增成功")

    def put(self):
        args = parser_department.parse_args()
        dept_id = args['dept_id']
        department_name = args['department_name']

        department = Department.find_by_id(dept_id)
        department.department_name = department_name
        Department.commit(self)
        return return_true_json("部门更新成功")

    def delete(self):
        args = parser_department.parse_args()
        dept_id = args['dept_id']

        department = Department.find_by_id(dept_id)
        Department.delete(department)
        return return_true_json("部门删除成功")


class ProductList(Resource):
    def get(self):
        datas = Product.find_all()
        res = [marshal(data, resource_product_fields) for data in datas]
        if datas:
            return return_true_json(res)
        else:
            return return_false_json(res)

    def post(self):
        args = parser_product.parse_args()
        product_name = args['product_name']
        product = Product(product_name=product_name)
        product.add_to_db()
        return return_true_json("产品新增成功")

    def put(self):
        args = parser_product.parse_args()
        pdt_id = args['pdt_id']
        product_name = args['product_name']

        product = Product.find_by_id(pdt_id)
        product.product_name = product_name
        Product.commit(self)
        return return_true_json("产品更新成功")

    def delete(self):
        args = parser_product.parse_args()
        pdt_id = args['pdt_id']

        product = Product.find_by_id(pdt_id)
        Product.delete(product)
        return return_true_json("产品删除成功")

class DictitemList(Resource):
    def get(self):
        datas = Dictitem.find_all()
        res = [marshal(data, resource_dictitem_fields) for data in datas]
        if datas:
            return return_true_json(res)
        else:
            return return_false_json(res)

#按照字典代码查询返回
class Dictitem_query(Resource):
    def get(self):
        dict_code = request.args['dict_code']
        datas = Dictitem.query_dict_code(dict_code)
        res = [marshal(datas, resource_dictitem_fields)]
        if datas:
            return return_true_json(res)
        else:
            return return_false_json(res)

    def post(self):
        args = parser_dictitem.parse_args()
        dict_code = args['dict_code']
        dict_name = args['dict_name']
        # key_value = json.dumps(args['key_value'],ensure_ascii=False)
        #将key_value获取的数组数据转成dict再重新生成数组序列
        key_value_old = args['key_value']
        key_value = []
        for i in range(len(key_value_old)):
            value = ast.literal_eval(key_value_old[i])
            key_value.append(value)
        key_value = json.dumps(key_value,ensure_ascii=False)
        print(key_value,type(key_value))

        dictitem = Dictitem(dict_code=dict_code,dict_name=dict_name,key_value=key_value)
        dictitem.add_to_db()
        return return_true_json("新增成功")

    def put(self):
        args = parser_dictitem.parse_args()
        id = args['id']
        dict_code = args['dict_code']
        dict_name = args['dict_name']
        key_value_old = args['key_value']
        key_value = []
        for i in range(len(key_value_old)):
            value = ast.literal_eval(key_value_old[i])
            key_value.append(value)
        key_value = json.dumps(key_value, ensure_ascii=False)

        dictitem = Dictitem.find_by_id(id)
        dictitem.dict_code = dict_code
        dictitem.dict_name = dict_name
        dictitem.key_value = key_value
        Dictitem.commit(self)
        return return_true_json("数据字典更新成功")

    def delete(self):
        args = parser_dictitem.parse_args()
        id = args['id']

        dictitem = Dictitem.find_by_id(id)
        Dictitem.delete(dictitem)
        return return_true_json("删除成功")

#导出周报写入到excel接口,根据数据自动判断本周完成和下周计划
class Weeklyreport_output(Resource):
    def post(self):
        args = parser_weeklyreport_output.parse_args()
        data = args["data"]
        new_data = []
        for i in range(len(data)):
            value = ast.literal_eval(data[i])
            new_data.append(value)
        # print(new_data)
        path = "D:\\2019周报.xlsx"
        data_finish = []
        data_plan = []
        # rwxz_value = Dictitem.query.filter(Dictitem.dict_code == 'rwxz').first().key_value #获取任务性质的字典值
        # l_rwxz_value = ast.literal_eval(rwxz_value)
        for i in range(len(new_data)):
            if new_data[i]['finished_time']:
                list =[
                    str(new_data[i]['create_time'][0:10].split("-")[1])+"月",
                    new_data[i]['create_time'][0:10],
                    new_data[i]['finished_time'][0:10],
                    "第"+str(get_week_of_month(new_data[i]['create_time']))+"周",
                    new_data[i]['task_type_name'],
                    new_data[i]['content'],
                    new_data[i]['product_name'],
                    new_data[i]['user_name'],
                    # new_data[i]['task_nature'],
                    get_dict_name("rwxz", new_data[i]['task_nature']),
                    100,
                    new_data[i]['finished_percent'],
                    new_data[i]['remark']
                ]
                data_finish.append(list)
            else:
                list = [
                    str(new_data[i]['create_time'][0:10].split("-")[1])+"月",
                    new_data[i]['create_time'][0:10],
                    new_data[i]['planfinished_time'][0:10],
                    "第"+str(get_week_of_month(new_data[i]['create_time']))+"周",
                    new_data[i]['task_type_name'],
                    new_data[i]['content'],
                    new_data[i]['product_name'],
                    new_data[i]['user_name'],
                    # new_data[i]['task_nature'],
                    get_dict_name("rwxz", new_data[i]['task_nature']),
                    new_data[i]['remark']
                ]
                data_plan.append(list)
        write_excel_xlsx(path, "本周完成", data_finish)
        write_excel_xlsx(path, "下周计划", data_plan)
        return return_true_json("xsxl表格追加数据成功")

#导出周报写入到excel接口
# class Weeklyreport_output(Resource):
#     def get(self):
#         args = parser_weeklyreport_output.parse_args()
#         path = args["path"]
#         sheet_name = args["sheet_name"]
#         value = args["value"]
#         value_new = []
#         for i in range(len(value)):
#             list = eval(value[i])
#             value_new.append(list)
#         print(value_new,type(value_new))
#         # creation_excel_xlsx(path, sheet_name, value_new)
#         # return return_true_json("xsxl表格生成成功")
#         write_excel_xlsx(path, sheet_name, value_new)
#         return return_true_json("xsxl表格追加数据成功")


#下载周报接口
class Weeklyreport_download(Resource):
    def get(self):
        args = parser_weeklyreport_download.parse_args()
        path = args["path"]
        if path is None:
            return return_false_json("没有路径参数")
        else:
            if path == '':
                return return_false_json("路径不能为空")
            else:
                if not os.path.isfile(path):
                    return return_false_json("文件路径不存在")
                else:
                    filename = os.path.basename(path)
                    print(filename)
                    response = Response(file_iterator(path))
                    print(response)
                    response.headers['Content-Type'] = 'application/octet-stream'
                    response.headers["Content-Disposition"] = 'attachment;filename="{}"'.format(filename)
                    print(response.headers)
                    return response